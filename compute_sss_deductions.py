from openpyxl import load_workbook
from functools import reduce

wb = load_workbook(filename='SSS-PAYROLL-2021-22-modified.xlsx')
wb_sss = load_workbook(filename='sss_2021_table.xlsx',data_only=True)

sss_ws = wb_sss['SSS 2021']
sss_starting_row = 7
sss_ending_row = sss_ws.max_row


_sss_starting_row = sss_starting_row
def get_sss_range_of_compensations(row):
    global _sss_starting_row
    row += (_sss_starting_row,)
    _sss_starting_row += 1
    return row





# function to get the sss deduction base on gross salary.
def get_sss_deduction(gross_salary, range_of_compensations):
    def get_sss_deduction_value(gross_salary, range):
        min_gross, max_gross, row = range
        if gross_salary >= int(min_gross.value) and gross_salary <= int(max_gross.value):
            return sss_ws[f'Q{row}'].value
    return reduce(
                lambda _, range :
                   get_sss_deduction_value(gross_salary, range), 
                range_of_compensations, 0
            )
            



# getting all available payroll monthly cutoff
excluded_payroll_sheets = ['MASTER LIST']
payroll_sheets = list(filter(lambda name : name not in excluded_payroll_sheets, wb.sheetnames))

range_of_compensations = [get_sss_range_of_compensations(row) for row in sss_ws.iter_rows(min_row=sss_starting_row,max_col=2)]


for payroll_month_year in payroll_sheets:
    payroll_ws = wb[payroll_month_year]
    payroll_month_year_starting_row = 6
    MAX_EMPLOYEE_ROWS = 34   # Don't forget to change there employee if needed

    for payroll_rows in payroll_ws['J{}:J{}'.format(payroll_month_year_starting_row,MAX_EMPLOYEE_ROWS)]:
        for payroll_cell in payroll_rows:
            value = payroll_cell.value
            if value > 0:
                payroll_ws[f'K{payroll_month_year_starting_row}'] = get_sss_deduction(value, range_of_compensations)
        payroll_month_year_starting_row += 1



wb.save('test.xlsx')
